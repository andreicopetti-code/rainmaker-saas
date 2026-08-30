import argparse
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill


def text_cell(worksheet, value, header=False):
    cell = WriteOnlyCell(worksheet, value="" if value is None else str(value))
    cell.data_type = "s"
    cell.number_format = "@"
    if header:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    return cell


def convert(source: Path, destination: Path):
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet("Empresas")
    worksheet.freeze_panes = "A2"

    row_count = 0
    with source.open("r", encoding="utf-8-sig", newline="") as csv_file:
        reader = csv.reader(csv_file, delimiter=";", quotechar='"')
        for row_count, row in enumerate(reader):
            worksheet.append(
                [text_cell(worksheet, value, header=row_count == 0) for value in row]
            )
            if row_count and row_count % 25000 == 0:
                print(f"{row_count:,} empresas gravadas...")

    workbook.save(destination)
    print(f"Concluído: {max(row_count, 0):,} empresas.")
    print(f"Arquivo: {destination}")


def main():
    parser = argparse.ArgumentParser(
        description="Converte o CSV tratado em XLSX sem alterar identificadores."
    )
    parser.add_argument("source", type=Path)
    parser.add_argument("destination", type=Path)
    args = parser.parse_args()
    convert(args.source.resolve(), args.destination.resolve())


if __name__ == "__main__":
    main()
